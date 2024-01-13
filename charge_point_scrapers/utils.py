import copy
import hashlib
import re
import os
import json

from typing import Union
from dateutil import parser
from scrapy import Request, Spider
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.utils import get_column_letter
from pathlib import PosixPath, Path
from dotenv import load_dotenv
from geopy import distance

from charge_point_scrapers.constants import ZAP_MAP_REQUEST_HEADERS, CO_CHARGER_REQUEST_HEADERS, SheetNames, EnvKeys

load_dotenv()


def copy_headers(original_headers, update_data: Union[dict, None]) -> dict:
    if not update_data:
        update_data = {}
    headers = copy.copy(original_headers)
    headers.update(**update_data)
    return headers


def request_zap_auth_token(callback):
    auth_req_headers = copy_headers(
        ZAP_MAP_REQUEST_HEADERS,
        {'Host': 'auth.zap-map.com'}
    )
    return Request(
        'https://auth.zap-map.com/guest/token',
        headers=auth_req_headers,
        callback=callback,
        dont_filter=True
    )


def parse_date(date_str: str):
    parsed_date = parser.parse(date_str)
    return parsed_date.strftime('%Y-%m-%d')


def remove_control_characters(cells):
    clean_strings = []
    for cell in cells:
        illegal_chars = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        if illegal_chars.findall(str(cell)):
            cell = re.sub(illegal_chars, ' ', str(cell))
        clean_strings.append(cell)
    return clean_strings


def filter_co_charger_radius(data):
    cities = {
        'Peterborough': (52.573552730184225, -0.25514820758860557),
        'Oxfordshire': (51.74383552545767, -1.2459381017248494),
        'Wimbledon': (51.42990367179369, -0.22438352969388295),
        'Bristol': (51.455539352076414, -2.5872982645856237),
        'Birmingham': (52.494308381894776, -1.8718849325622116),
        'Manchester': (53.48602564820113, -2.2399557350492376),
        'Milton Keynes': (52.040187985429135, -0.7580044689136244),
        'Reading': (51.45420910597151, -0.9802584269217173)
    }
    matching_hosts = {}
    for host in data:
        for city, city_cords in cities.items():
            geo_dist = distance.distance(city_cords, (float(host['latitude']), float(host['longitude'])))
            if geo_dist.miles < 21.0:
                if host['id'] in matching_hosts:
                    continue
                else:
                    matching_hosts[host['id']] = host

    return matching_hosts


def export_to_excel(
        excel_file_path: Union[PosixPath, Path],
        jsonl_feed_path: Union[PosixPath, Path],
        col_mapping: dict,
        sheet_name: str,
        spider: Spider
):
    """
    Export json lines feed to excel.
    IF sheet name exists, delete
    Always creates a new sheet with the name
    """
    if excel_file_path.exists():
        workbook = load_workbook(filename=excel_file_path)
    else:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)

    sheet = workbook.create_sheet(sheet_name, 0)

    column_lens = {}

    for col_num, value in enumerate(col_mapping.values(), start=1):
        cell = sheet.cell(row=1, column=col_num, value=value)
        cell.font = Font(bold=True, size=12)
        column_letter = get_column_letter(col_num)
        column_lens[column_letter] = len(value) * 1.2
        sheet.column_dimensions[column_letter].width = column_lens[column_letter]

    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, size=14)
    data = []
    try:
        with open(jsonl_feed_path, 'r', encoding='utf-8') as f:
            for point in f.readlines():
                data.append(json.loads(point))
    except FileNotFoundError as ex:
        spider.logger.exception(ex)
        return

    ignore_radius_filter = os.getenv(EnvKeys.CO_CHARGER_IGNORE_20_MILE_RADIUS.value, 0)
    if sheet_name == SheetNames.CO_CHARGER.value and not ignore_radius_filter:
        data = filter_co_charger_radius(data)

    for row in data:
        row_val = []
        for idx, key in enumerate(col_mapping, start=1):
            val = row[key]
            if key in ['address', 'street']:
                val = val.replace('\r', '')
                val = val.replace('\n', ' ')
            if not val:
                val = 'N/A'
            column_letter = get_column_letter(idx)
            if len(val) > column_lens.get(column_letter, 0):
                if key in ['address', 'name']:
                    column_lens[column_letter] = len(val) * .8
                elif key in ['charging_fee', 'parking_fee']:
                    column_lens[column_letter] = len(val) * .53
                else:
                    column_lens[column_letter] = len(val)
                sheet.column_dimensions[column_letter].width = column_lens[column_letter]

            row_val.append(val)

        try:
            sheet.append(row_val)
        except IllegalCharacterError as ex:
            sheet_row = remove_control_characters(row_val)
            sheet.append(sheet_row)

    workbook.save(filename=excel_file_path)


def authenticate_co_charger(callback):
    co_charger_email = os.getenv(EnvKeys.CO_CHARGER_EMAIL.value)
    co_charger_password = os.getenv(EnvKeys.CO_CHARGER_PASSWORD.value)

    if not co_charger_email or not co_charger_password:
        raise Exception('Username or password not set. Update the values in .env file')

    login_payload = {
        "email": co_charger_email,
        "password": hashlib.md5(co_charger_password.encode('utf-8')).hexdigest(),
        "platform": "Platform: android - OS Version: 13 - Device: M2101K6G"
    }

    return Request(
        'https://api.co-charger.com/login',
        method='POST',
        headers=CO_CHARGER_REQUEST_HEADERS,
        body=json.dumps(login_payload),
        callback=callback
    )


def fmt_co_charger_value(raw_value):
    if not raw_value or raw_value.strip() == 'null':
        return ''

    return raw_value.strip() or ''


def update_co_charger_auth_token(token):
    root_dir = Path(__file__).parent.parent
    current_lines = [
        f"{env_key.name}={os.getenv(env_key.value)}\n" for env_key in EnvKeys
        if env_key.name != EnvKeys.CO_CHARGER_TOKEN.name
    ]
    current_lines.append(f'{EnvKeys.CO_CHARGER_TOKEN.name}={token}\n')

    with open(root_dir / '.env', 'w+', encoding='utf-8') as f:
        for line in current_lines:
            f.write(line)
